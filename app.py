import os
from flask import Flask, render_template, request, send_file, jsonify, abort
from markupsafe import Markup
import requests
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment  # Добавим импорт в начало файла
from openpyxl.worksheet.hyperlink import Hyperlink  # Обновленный импорт
from datetime import datetime, timedelta
import json
import threading

app = Flask(__name__)

print(os.getcwd())

# Загрузка токена из файла конфигурации
with open('config.py', 'r') as config_file:
    exec(config_file.read())

BASE_URL = 'https://api.moysklad.ru/api/remap/1.2'

# Добавим глобальную переменную для отслеживания состояния
processing_cancelled = False
processing_lock = threading.Lock()

def render_group_options(groups, level=0):
    result = []
    for group in groups:
        indent = '—' * level
        result.append(f'<option value="{group["id"]}" {"class=\"nested-select\"" if level > 0 else ""} style="margin-left: {level * 20}px;">{indent} {group["name"]}</option>')
        if group.get('children'):
            result.extend(render_group_options(group['children'], level + 1))
    return '\n'.join(result)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        store_id = request.form['store_id']
        planning_days = int(request.form['planning_days'])  # Получаем количетво дней
        product_groups = request.form['product_group'].split(',') if request.form['product_group'] else []
        
        print(f"Received form data: start_date={start_date}, end_date={end_date}, store_id={store_id}, product_groups={product_groups}, planning_days={planning_days}")
        
        try:
            report_data = get_report_data(start_date, end_date, store_id, product_groups)
            
            if not report_data or 'rows' not in report_data or not report_data['rows']:
                return "Нет данных для формирования отчета для выбранных параметров", 404
            
            excel_file = create_excel_report(report_data, store_id, end_date, planning_days)  # Передаем planning_days
            
            return send_file(excel_file, as_attachment=True, download_name='profitability_report.xlsx')
        except Exception as e:
            print(f"Error in index(): {str(e)}")
            return f"Произошла ошибка при формировании отчета: {str(e)}", 500
    
    stores = get_stores()
    product_groups = get_product_groups()
    return render_template('index.html', stores=stores, product_groups=product_groups, render_group_options=render_group_options)

@app.route('/get_subgroups/<group_id>')
def get_subgroups(group_id):
    subgroups = get_subgroups_for_group(group_id)
    return jsonify(subgroups)

@app.route('/stop_processing', methods=['POST'])
def stop_processing():
    global processing_cancelled
    with processing_lock:
        processing_cancelled = True
    return '', 204

def check_if_cancelled():
    global processing_cancelled
    with processing_lock:
        if processing_cancelled:
            raise Exception("Processing cancelled by user")

def get_report_data(start_date, end_date, store_id, product_groups):
    global processing_cancelled
    with processing_lock:
        processing_cancelled = False
    
    try:
        url = f"{BASE_URL}/report/profit/byvariant"
        headers = {
            'Authorization': f'Bearer {MOYSKLAD_TOKEN}',
            'Accept': 'application/json;charset=utf-8',
            'Content-Type': 'application/json'
        }
        
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
        
        formatted_start = start_datetime.strftime('%Y-%m-%d %H:%M:%S')
        formatted_end = end_datetime.replace(hour=23, minute=59, second=59).strftime('%Y-%m-%d %H:%M:%S')
        
        params = {
            'momentFrom': formatted_start,
            'momentTo': formatted_end,
            'limit': 1000,
            'offset': 0
        }
        
        filter_parts = []
        
        if store_id:
            store_url = f"{BASE_URL}/entity/store/{store_id}"
            filter_parts.append(f'store={store_url}')
        
        # Modified to handle multiple product groups
        if product_groups:
            for group_id in product_groups:
                if group_id:  # Check if group_id is not empty
                    product_folder_url = f"{BASE_URL}/entity/productfolder/{group_id}"
                    filter_parts.append(f'productFolder={product_folder_url}')
        
        if filter_parts:
            # Join all filter parts with semicolon
            params['filter'] = ';'.join(filter_parts)
        
        all_rows = []
        total_count = None
        
        while True:
            check_if_cancelled()  # Проверяем отмену
            
            # Формируем строку запроса
            query_params = [f"{k}={v}" for k, v in params.items() if k != 'filter']
            if 'filter' in params:
                query_params.append(f"filter={params['filter']}")
            query_string = '&'.join(query_params)
            
            full_url = f"{url}?{query_string}"
            print(f"Отправляем запрос: URL={full_url}, Headers={headers}")
            
            response = requests.get(full_url, headers=headers)
            
            if response.status_code != 200:
                error_message = f"Ошибка при получении данных: {response.status_code}. Ответ сервера: {response.text}"
                print(error_message)
                raise Exception(error_message)
            
            data = response.json()
            
            if total_count is None:
                total_count = data.get('meta', {}).get('size', 0)
                print(f"Всего записей: {total_count}")
            
            all_rows.extend(data.get('rows', []))
            
            if len(all_rows) >= total_count:
                break
            
            params['offset'] += params['limit']
        
        return {'meta': data.get('meta', {}), 'rows': all_rows}
        
    except Exception as e:
        if str(e) == "Processing cancelled by user":
            abort(499, description="Processing cancelled by user")
        raise e

def get_stores():
    url = f"{BASE_URL}/entity/store"
    headers = {
        'Authorization': f'Bearer {MOYSKLAD_TOKEN}',
        'Accept': 'application/json;charset=utf-8'
    }
    
    print(f"Отправляем запрос для получения списка складов: URL={url}, Headers={headers}")  # Для отладки
    
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        stores = response.json()['rows']
        return [{'id': store['id'], 'name': store['name']} for store in stores]
    else:
        error_message = f"Ошибка при получении списка складов: {response.status_code}. Ответ сервера: {response.text}"
        print(error_message)  # Выводим ошибку в консоль для отладки
        raise Exception(error_message)
    
def get_subgroups_for_group(group_id):
    url = f"{BASE_URL}/entity/productfolder"
    headers = {
        'Authorization': f'Bearer {MOYSKLAD_TOKEN}',
        'Accept': 'application/json;charset=utf-8'
    }
    params = {
        'filter': f'productFolder={group_id}'
    }
    
    response = requests.get(url, headers=headers, params=params)
    if response.status_code == 200:
        subgroups = response.json()['rows']
        return [{'id': group['id'], 'name': group['name'], 'children': []} for group in subgroups]
    else:
        error_message = f"Ошибка при получении подгрупп: {response.status_code}. Ответ сервера: {response.text}"
        print(error_message)
        return []

def get_product_groups():
    url = f"{BASE_URL}/entity/productfolder"
    headers = {
        'Authorization': f'Bearer {MOYSKLAD_TOKEN}',
        'Accept': 'application/json;charset=utf-8'
    }
    
    all_groups = []
    offset = 0
    limit = 1000

    while True:
        params = {
            'offset': offset,
            'limit': limit
        }
        
        response = requests.get(url, headers=headers, params=params)
        if response.status_code == 200:
            data = response.json()
            all_groups.extend(data['rows'])
            if len(data['rows']) < limit:
                break
            offset += limit
        else:
            error_message = f"Ошибка при получении ска групп товаров: {response.status_code}. Ответ сервера: {response.text}"
            print(error_message)
            raise Exception(error_message)

    return build_group_hierarchy(all_groups)

def build_group_hierarchy(groups):
    group_dict = {}
    root_groups = []

    # Первый проход: создаем словарь всех групп
    for group in groups:
        group_id = group['id']
        group_dict[group_id] = {
            'id': group_id,
            'name': group['name'],
            'children': [],
            'parent': None
        }

    # Второй проход: устанавливаем связи родитель-потомок
    for group in groups:
        group_id = group['id']
        parent_href = group.get('productFolder', {}).get('meta', {}).get('href')
        if parent_href:
            parent_id = parent_href.split('/')[-1]
            if parent_id in group_dict:
                group_dict[group_id]['parent'] = parent_id
                group_dict[parent_id]['children'].append(group_dict[group_id])
        else:
            root_groups.append(group_dict[group_id])

    # Сртируем группы по имени
    root_groups.sort(key=lambda x: x['name'])
    for group in group_dict.values():
        group['children'].sort(key=lambda x: x['name'])

    return root_groups

# Добавьте эту функцию для отладки
def print_group_hierarchy(groups, level=0):
    for group in groups:
        print("  " * level + f"{group['name']} (ID: {group['id']})")
        print_group_hierarchy(group['children'], level + 1)

def get_sales_speed(variant_id, store_id, end_date, is_variant):
    url = f"{BASE_URL}/report/turnover/byoperations"
    headers = {
        'Authorization': f'Bearer {MOYSKLAD_TOKEN}',
        'Accept': 'application/json;charset=utf-8'
    }
    
    start_date = "2024-01-01 00:00:00"
    end_date_formatted = datetime.strptime(end_date, '%Y-%m-%d').strftime('%Y-%m-%d 23:59:59')
    
    assortment_type = 'variant' if is_variant else 'product'
    
    filter_params = [
        f"filter=store={BASE_URL}/entity/store/{store_id}",
        f"filter={assortment_type}={BASE_URL}/entity/{assortment_type}/{variant_id}"
    ]
    
    params = {
        'momentFrom': start_date,
        'momentTo': end_date_formatted,
    }
    
    query_string = '&'.join([f"{k}={v}" for k, v in params.items()] + filter_params)
    full_url = f"{url}?{query_string}"
    
    print(f"Запрос для получения данных о продажах: URL={full_url}")
    
    response = requests.get(full_url, headers=headers)
    if response.status_code != 200:
        print(f"Ошибка пи получении данных о проажах: {response.status_code}. Ответ ервера: {response.text}")
        return 0, '', '', '', ''  # Возвращаем 0 для скорости и пустую строку для UUID

    data = response.json()
    rows = data.get('rows', [])

    # Получаем UUID группы и название группы из первой строки данных
    group_uuid = ''
    group_name = ''
    if rows:
        assortment = rows[0].get('assortment', {})
        product_folder = assortment.get('productFolder', {})
        group_href = product_folder.get('meta', {}).get('href', '')
        group_uuid = group_href.split('/')[-1] if group_href else ''
        group_name = product_folder.get('name', '')  # Получаем название группы
        print(f"Found group UUID: {group_uuid}, name: {group_name}")

    # олучаем UUID и сылку на товар из первой строки данных
    product_uuid = ''
    product_href = ''
    if rows:
        assortment = rows[0].get('assortment', {})
        product_meta = assortment.get('meta', {})
        product_href = product_meta.get('uuidHref', '')
        if product_href:
            product_uuid = product_href.split('/')[-1]

    # Фильтрация по UUID модификации
    filtered_rows = [
        row for row in rows
        if row.get('assortment', {}).get('meta', {}).get('href', '').split('/')[-1] == variant_id
    ]

    # Сортировка операций по дате
    filtered_rows.sort(key=lambda x: datetime.fromisoformat(x['operation']['moment'].replace('Z', '+00:00')))

    retail_demand_counter = 0
    current_stock = 0
    last_operation_time = None
    on_stock_time = timedelta()

    end_datetime = datetime.strptime(end_date_formatted, '%Y-%m-%d %H:%M:%S')

    for row in filtered_rows:
        quantity = row['quantity']
        operation_time = datetime.fromisoformat(row['operation']['moment'].replace('Z', '+00:00'))
        operation_type = row['operation']['meta']['type']

        if last_operation_time and current_stock > 0:
            on_stock_time += operation_time - last_operation_time

        if quantity > 0:  # Приход товара
            current_stock += quantity
        else:  # Уход товара
            quantity = abs(quantity)
            current_stock = max(0, current_stock - quantity)
            if operation_type == 'retaildemand':
                retail_demand_counter += quantity

        last_operation_time = operation_time

    # Учитываем время от последней операции до конца периода
    if last_operation_time and current_stock > 0:
        on_stock_time += end_datetime - last_operation_time

    days_on_stock = on_stock_time.total_seconds() / (24 * 60 * 60)

    if days_on_stock > 0:
        sales_speed = round(retail_demand_counter / days_on_stock, 2)
    else:
        sales_speed = 0

    print(f"sales_speed: {sales_speed}, group_uuid: {group_uuid}, product_href: {product_href}")

    return sales_speed, group_uuid, group_name, product_uuid, product_href

def get_group_path(group_uuid, product_groups, get_uuid=False):
    def find_group_path(groups, target_uuid, current_path=[], current_uuid_path=[]):
        for group in groups:
            if group['id'] == target_uuid:
                return (current_path + [group['name']], current_uuid_path + [group['id']])
            if group.get('children'):
                path = find_group_path(group['children'], target_uuid, 
                                     current_path + [group['name']], 
                                     current_uuid_path + [group['id']])
                if path:
                    return path
        return None

    path = find_group_path(product_groups, group_uuid)
    if not path:
        return '', []  # Возвращаем пустую строку и пустой список UUID
    
    names_path, uuid_path = path
    return ('/'.join(names_path), uuid_path) if not get_uuid else ('/'.join(uuid_path), uuid_path)

def get_sheet_name(products_data):
    # Получаем уникальные названи второго уровня
    level2_names = set()
    for product in products_data:
        names_by_level = product['names_by_level']
        if len(names_by_level) > 1:  # Есл есь второй уровень
            level2_names.add(names_by_level[1])
    
    # Сортируем имена для консистентности
    level2_names = sorted(list(level2_names))
    
    # Формируем название листа с новым разделителем
    if len(level2_names) > 5:
        sheet_name = ', '.join(level2_names[:5])
    else:
        sheet_name = ', '.join(level2_names)
    
    # Ограничиваем длину названия листа (максимум 31 символ в Excel)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:28] + "..."
    
    # Если название пустое, используем значение по умолчанию
    return sheet_name if sheet_name else "Отчет приб��льности"

def create_excel_report(data, store_id, end_date, planning_days):
    try:
        print("Начало создания Excel отчета")
        wb = Workbook()
        ws = wb.active
        
        product_groups = get_product_groups()
        products_data = []
        max_depth = 0
        
        # Сначала собираем все данные и определяем максимальную глубину
        for item in data['rows']:
            check_if_cancelled()
            assortment = item.get('assortment', {})
            assortment_meta = assortment.get('meta', {})
            assortment_href = assortment_meta.get('href', '')
            
            is_variant = '/variant/' in assortment_href
            variant_id = assortment_href.split('/variant/')[-1] if is_variant else assortment_href.split('/product/')[-1]
            
            if variant_id:
                sales_speed, group_uuid, group_name, product_uuid, product_href = get_sales_speed(variant_id, store_id, end_date, is_variant)
                if sales_speed != 0:
                    full_path, uuid_path = get_group_path(group_uuid, product_groups)
                    max_depth = max(max_depth, len(uuid_path))  # Используем длину списка UUID
                    
                    products_data.append({
                        'name': assortment.get('name', ''),
                        'quantity': item.get('sellQuantity', 0),
                        'profit': round(item.get('profit', 0) / 100, 2),
                        'sales_speed': sales_speed,
                        'forecast': sales_speed * planning_days,
                        'group_uuid': group_uuid,
                        'group_path': full_path,
                        'uuid_path': uuid_path,  # Сохраняем список UUID для правильного определения уровней
                        'names_by_level': get_names_by_uuid(uuid_path, product_groups),
                        'product_uuid': product_uuid,
                        'product_href': product_href
                    })

        print(f"Максимальная глубина групп: {max_depth}")

        # Сортируем данные по полному пути групп по возрастанию
        products_data.sort(key=lambda x: x['group_path'])

        # Формируем заголовки с учетом реальной глубины, начиная со второго уровня
        group_level_headers = [f'Уровень {i+2}' for i in range(max_depth-1)] if max_depth > 1 else []
        headers = group_level_headers + [
            'UUID',  # Изменено название столбца
            'Наименоване', 'Количество', 'Прибыльность', 'Скорость продаж', 
            f'Прогноз на {planning_days} дней'
        ]
        
        # Записываем заголовки
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        # Записываем данные с группами
        current_row = 2
        current_uuid_path = []
        
        for product in products_data:
            uuid_path = product['uuid_path']
            names_by_level = product['names_by_level']
            
            # Записываем строки групп, если путь изменился
            for i, uuid in enumerate(uuid_path):
                if i >= len(current_uuid_path) or uuid != current_uuid_path[i]:
                    if i > 0:
                        ws.cell(row=current_row, column=i, value=names_by_level[i])
                    # Записываем UUID группы полностью
                    uuid_cell = ws.cell(row=current_row, column=max_depth, value=uuid)
                    uuid_cell.alignment = Alignment(horizontal='left', shrink_to_fit=False)
                    current_row += 1
            
            # При записи UUID товара
            uuid_cell = ws.cell(row=current_row, column=max_depth)
            if product['product_href']:
                uuid_cell.value = product['product_uuid']  # Записываем полный UUID
                uuid_cell.hyperlink = product['product_href']
                uuid_cell.font = Font(color="0000FF", underline="single")
                uuid_cell.alignment = Alignment(horizontal='left', shrink_to_fit=False)
            
            ws.cell(row=current_row, column=max_depth+1, value=product['name'])
            ws.cell(row=current_row, column=max_depth+2, value=product['quantity'])
            ws.cell(row=current_row, column=max_depth+3, value=product['profit'])
            ws.cell(row=current_row, column=max_depth+4, value=product['sales_speed'])
            ws.cell(row=current_row, column=max_depth+5, value=product['forecast'])
            
            current_row += 1
            current_uuid_path = uuid_path

        # Обновляем диапазон таблицы с учетом реальной глубины
        last_col = max_depth + 5  # 5 - количество основных столбцов с учетом нового столбца UUID
        table_ref = f"A1:{chr(64 + last_col)}{current_row-1}"
        tab = Table(displayName="Table1", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # Форматирование
        ws.freeze_panes = 'A2'
        
        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            # Если это столбец "UUID" (max_depth)
            if column[0].column == max_depth:
                ws.column_dimensions[column_letter].width = 3
                # Применяем настройки отображения ко всем ячейкам в столбце
                for cell in column:
                    if isinstance(cell.hyperlink, str):  # Если есть ссылка
                        cell.font = Font(color="0000FF", underline="single")
                    cell.alignment = Alignment(horizontal='left', shrink_to_fit=False)
                continue
                
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # После сбора всех данных и перед созданием заголовков
        sheet_name = get_sheet_name(products_data)
        ws.title = sheet_name
        print(f"Название листа: {sheet_name}")
        
        filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb.save(filename)
        wb.close()
        return filename
        
    except Exception as e:
        if str(e) == "Processing cancelled by user":
            abort(499, description="Processing cancelled by user")
        raise e
    finally:
        try:
            wb.close()
        except:
            pass

def get_names_by_uuid(uuid_path, product_groups):
    def find_name_by_uuid(groups, target_uuid):
        for group in groups:
            if group['id'] == target_uuid:
                return group['name']
            if group.get('children'):
                name = find_name_by_uuid(group['children'], target_uuid)
                if name:
                    return name
        return None

    return [find_name_by_uuid(product_groups, uuid) or '' for uuid in uuid_path]

if __name__ == '__main__':
    app.run(debug=True)
